#!/usr/bin/env python3
"""
Q2 Reporting Agent
Usage: python q2_agent.py <excel_file> [--output <output_dir>]
"""

import argparse
import sys
import os
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Parsers per sheet
# ---------------------------------------------------------------------------

def parse_trial_balance(wb):
    ws = wb["Trial Balance"]
    rows = list(ws.iter_rows(values_only=True))

    q1_label = q2_label = None
    for row in rows[:4]:
        for cell in row:
            if isinstance(cell, str) and "Period" in cell:
                if q1_label is None:
                    q1_label = cell.replace("Period = ", "")
                else:
                    q2_label = cell.replace("Period = ", "")

    # Columns: Q1 = cols 0-5, Q2 = cols 7-12
    # Row 6 onward = data rows (index 6)
    accounts = []
    for row in rows[6:]:
        code = row[0]
        name = row[1]
        q1_end = row[5]
        q2_end = row[12]
        if code and name and isinstance(q1_end, (int, float)):
            accounts.append({
                "code": code,
                "name": name,
                "q1_ending": float(q1_end) if q1_end is not None else 0.0,
                "q2_ending": float(q2_end) if q2_end is not None else 0.0,
            })

    return {"q1_label": q1_label, "q2_label": q2_label, "accounts": accounts}


def parse_banken(wb):
    ws = wb["Banken"]
    banks = []
    total = 0.0
    for row in ws.iter_rows(values_only=True):
        code = row[1]
        name = row[2]
        balance = row[3]
        if code and name and isinstance(balance, (int, float)):
            banks.append({"code": code, "name": name, "balance": float(balance)})
        if name and "Liquide middelen" in str(name) and isinstance(balance, (int, float)):
            total = float(balance)
    return {"banks": banks, "total": total}


def parse_ar(wb):
    ws = wb["AR"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        if row[0] == "Legal":
            header_row = i
            break

    if header_row is None:
        return {"rows": [], "total_owed": 0.0, "overdue": 0.0}

    data_rows = []
    for row in rows[header_row + 2:]:
        legal = row[0]
        prop = row[1]
        charge_to = row[2]
        total_charged = row[4]
        current_owed = row[6]
        over_150 = row[12]
        total_owed = row[14]

        if charge_to and isinstance(total_owed, (int, float)) and total_owed != 0:
            data_rows.append({
                "property": prop or "",
                "tenant": charge_to,
                "total_charged": float(total_charged) if isinstance(total_charged, (int, float)) else 0.0,
                "current_owed": float(current_owed) if isinstance(current_owed, (int, float)) else 0.0,
                "over_150": float(over_150) if isinstance(over_150, (int, float)) else 0.0,
                "total_owed": float(total_owed),
            })

    total_owed = sum(r["total_owed"] for r in data_rows)
    overdue = sum(r["over_150"] for r in data_rows)
    return {"rows": data_rows, "total_owed": total_owed, "overdue": overdue}


def parse_ap(wb):
    ws = wb["AP"]
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for i, row in enumerate(rows):
        if row[0] == "Payee" and row[1] == "Payee Name":
            header_row = i
            break

    if header_row is None:
        return {"rows": [], "total": 0.0}

    data_rows = []
    for row in rows[header_row + 2:]:
        code = row[0]
        name = row[1]
        current = row[3]
        over_90 = row[7]
        if name and isinstance(current, (int, float)):
            total = sum(
                float(row[i]) if isinstance(row[i], (int, float)) else 0.0
                for i in range(3, 9)
            )
            data_rows.append({
                "code": code or "",
                "name": name,
                "current": float(current) if isinstance(current, (int, float)) else 0.0,
                "over_90": float(over_90) if isinstance(over_90, (int, float)) else 0.0,
                "total": total,
            })

    total = sum(r["total"] for r in data_rows)
    return {"rows": data_rows, "total": total}


def parse_periodieke_huur(wb):
    ws = wb["Periodieke huur"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains datetime objects as column headers for months)
    header_row = None
    for i, row in enumerate(rows):
        if row[0] == "Timeless" and row[3] == "Unit":
            header_row = i
            break

    if header_row is None:
        return {"tenants": [], "q2_total": 0.0, "months": []}

    header = rows[header_row]
    month_cols = []
    for j, cell in enumerate(header):
        if isinstance(cell, datetime) and cell.month in (4, 5, 6):
            month_cols.append((j, cell))

    tenants = []
    for row in rows[header_row + 1:]:
        timeless = row[0]
        omschrijving = row[1]
        prop = row[2]
        unit = row[3]
        tenant = row[4]

        if not tenant or str(tenant).startswith("Totaal"):
            continue

        monthly = {}
        for col_idx, month_dt in month_cols:
            val = row[col_idx]
            monthly[month_dt.strftime("%b %Y")] = float(val) if isinstance(val, (int, float)) else 0.0

        q2_total = sum(monthly.values())
        if q2_total != 0:
            tenants.append({
                "account": timeless or "",
                "description": omschrijving or "",
                "property": prop or "",
                "unit": unit or "",
                "tenant": tenant,
                "monthly": monthly,
                "q2_total": q2_total,
            })

    months = [m.strftime("%b %Y") for _, m in month_cols]
    q2_total = sum(t["q2_total"] for t in tenants)
    return {"tenants": tenants, "q2_total": q2_total, "months": months}


def parse_leningen(wb):
    ws = wb["Leningen - Rente"]
    rows = list(ws.iter_rows(values_only=True))

    loans = []
    interest_total = 0.0
    for row in rows[2:]:
        code = row[0]
        desc = row[1]
        balance = row[2]
        if code and desc and isinstance(balance, (int, float)):
            if str(code).startswith("2205"):
                loans.append({"code": code, "name": desc, "balance": float(balance)})
            elif str(code).startswith("8204"):
                interest_total += float(row[13]) if isinstance(row[13], (int, float)) else 0.0

    total_debt = sum(l["balance"] for l in loans)
    return {"loans": loans, "total_debt": total_debt, "interest_total": interest_total}


def parse_waarborgsommen(wb):
    ws = wb["Waarborgsommen"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        if row[0] == "Property" and row[1] == "Property Name":
            header_row = i
            break

    if header_row is None:
        return {"total": 0.0, "count": 0}

    total = 0.0
    count = 0
    for row in rows[header_row + 1:]:
        mutaties = row[9]
        if isinstance(mutaties, (int, float)):
            total += float(mutaties)
            count += 1

    return {"total": abs(total), "count": count}


def parse_tenancy_schedule(wb):
    ws = wb["Tenancy Schedule II "]
    rows = list(ws.iter_rows(values_only=True))

    # Header is multi-row; data starts around row 5
    leases = []
    for row in rows[5:]:
        prop = row[0]
        unit = row[14]
        lease_from = row[3]
        lease_to = row[4]
        monthly_rent = row[18]
        annual_rent = row[8]
        tenant = row[24]

        if unit and isinstance(monthly_rent, (int, float)) and monthly_rent > 0:
            leases.append({
                "property": str(prop).strip() if prop else "",
                "unit": unit,
                "tenant": str(tenant).strip() if tenant else "",
                "lease_from": lease_from.strftime("%d-%m-%Y") if isinstance(lease_from, datetime) else "",
                "lease_to": lease_to.strftime("%d-%m-%Y") if isinstance(lease_to, datetime) else "Onbepaald",
                "monthly_rent": float(monthly_rent),
                "annual_rent": float(annual_rent) if isinstance(annual_rent, (int, float)) else 0.0,
            })

    active = [l for l in leases if not l["lease_to"] or l["lease_to"] == "Onbepaald"]
    return {"leases": leases, "total_units": len(leases), "total_annual_rent": sum(l["annual_rent"] for l in leases)}


def parse_btw(wb):
    ws = wb["BTW"]
    rows = list(ws.iter_rows(values_only=True))

    q2_totals = {}
    for row in rows[5:]:
        label = row[1]
        q2_val = row[6]
        if label and isinstance(q2_val, (int, float)):
            q2_totals[str(label)] = float(q2_val)

    return {"q2": q2_totals}


# ---------------------------------------------------------------------------
# HTML report generation
# ---------------------------------------------------------------------------

REPORT_TEMPLATE = """<!DOCTYPE html>
<html lang="nl">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Q2 Rapportage – {company}</title>
<style>
  :root {{
    --primary: #1a3a5c;
    --accent: #2e86ab;
    --light: #f4f7fb;
    --border: #dce3ed;
    --green: #27ae60;
    --red: #e74c3c;
    --orange: #e67e22;
    --text: #2c3e50;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: var(--light); color: var(--text); font-size: 14px; }}

  .page-header {{
    background: var(--primary);
    color: #fff;
    padding: 32px 40px 24px;
  }}
  .page-header h1 {{ font-size: 26px; font-weight: 700; }}
  .page-header .meta {{ margin-top: 6px; font-size: 13px; opacity: .75; }}

  .container {{ max-width: 1200px; margin: 0 auto; padding: 32px 24px; }}

  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 32px;
  }}
  .kpi {{
    background: #fff;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 20px 18px;
    border-top: 4px solid var(--accent);
  }}
  .kpi .label {{ font-size: 11px; text-transform: uppercase; letter-spacing: .6px; color: #6b7d93; margin-bottom: 6px; }}
  .kpi .value {{ font-size: 22px; font-weight: 700; color: var(--primary); }}
  .kpi .sub {{ font-size: 11px; color: #8a9bb0; margin-top: 4px; }}

  .section {{ margin-bottom: 36px; }}
  .section-title {{
    font-size: 15px;
    font-weight: 700;
    color: var(--primary);
    padding: 10px 0 8px;
    border-bottom: 2px solid var(--accent);
    margin-bottom: 14px;
  }}

  table {{ width: 100%; border-collapse: collapse; background: #fff; border-radius: 8px; overflow: hidden; border: 1px solid var(--border); }}
  th {{ background: var(--primary); color: #fff; padding: 10px 12px; font-size: 12px; text-align: left; font-weight: 600; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid var(--border); font-size: 13px; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: #f0f4f9; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .pos {{ color: var(--green); }}
  .neg {{ color: var(--red); }}
  .warn {{ color: var(--orange); }}

  .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  @media (max-width: 768px) {{ .two-col {{ grid-template-columns: 1fr; }} }}

  .footer {{ text-align: center; font-size: 11px; color: #aaa; padding: 24px 0 8px; }}
</style>
</head>
<body>

<div class="page-header">
  <h1>Q2 Rapportage &mdash; {company}</h1>
  <div class="meta">Periode: {period} &nbsp;|&nbsp; Gegenereerd op: {generated}</div>
</div>

<div class="container">

  <!-- KPI row -->
  <div class="kpi-grid">
    <div class="kpi">
      <div class="label">Huurinkomsten Q2</div>
      <div class="value">{rent_q2}</div>
      <div class="sub">Periodieke huur</div>
    </div>
    <div class="kpi">
      <div class="label">Bankstand</div>
      <div class="value">{bank_total}</div>
      <div class="sub">Liquide middelen</div>
    </div>
    <div class="kpi">
      <div class="label">Debiteuren (totaal)</div>
      <div class="value">{ar_total}</div>
      <div class="sub">AR openstaand</div>
    </div>
    <div class="kpi">
      <div class="label">Crediteuren</div>
      <div class="value">{ap_total}</div>
      <div class="sub">AP openstaand</div>
    </div>
    <div class="kpi">
      <div class="label">Hypotheekschuld</div>
      <div class="value">{debt_total}</div>
      <div class="sub">Rabobank</div>
    </div>
    <div class="kpi">
      <div class="label">Waarborgsommen</div>
      <div class="value">{deposit_total}</div>
      <div class="sub">Ontvangen van huurders</div>
    </div>
  </div>

  <!-- Rent income -->
  <div class="section">
    <div class="section-title">Huurinkomsten Q2 per huurder</div>
    <table>
      <thead>
        <tr>
          <th>Unit</th>
          <th>Huurder</th>
          <th>Rekening</th>
          {rent_month_headers}
          <th class="num">Q2 Totaal</th>
        </tr>
      </thead>
      <tbody>
        {rent_rows}
        <tr style="font-weight:700; background:#eef2f8;">
          <td colspan="3">Totaal</td>
          {rent_total_months}
          <td class="num">{rent_q2_fmt}</td>
        </tr>
      </tbody>
    </table>
  </div>

  <!-- AR & AP -->
  <div class="two-col">
    <div class="section">
      <div class="section-title">Debiteuren (AR Aging)</div>
      <table>
        <thead>
          <tr>
            <th>Huurder</th>
            <th class="num">Huidig</th>
            <th class="num">Totaal open</th>
          </tr>
        </thead>
        <tbody>
          {ar_rows}
        </tbody>
      </table>
    </div>

    <div class="section">
      <div class="section-title">Crediteuren (AP Aging)</div>
      <table>
        <thead>
          <tr>
            <th>Naam</th>
            <th class="num">Huidig</th>
            <th class="num">Totaal open</th>
          </tr>
        </thead>
        <tbody>
          {ap_rows}
        </tbody>
      </table>
    </div>
  </div>

  <!-- Banks -->
  <div class="section">
    <div class="section-title">Bankrekeningen</div>
    <table>
      <thead>
        <tr>
          <th>Rekening</th>
          <th>Omschrijving</th>
          <th class="num">Saldo</th>
        </tr>
      </thead>
      <tbody>
        {bank_rows}
        <tr style="font-weight:700; background:#eef2f8;">
          <td colspan="2">Totaal liquide middelen</td>
          <td class="num">{bank_total_fmt}</td>
        </tr>
      </tbody>
    </table>
  </div>

  <!-- Loans -->
  <div class="section">
    <div class="section-title">Leningen &amp; Rente</div>
    <table>
      <thead>
        <tr>
          <th>Code</th>
          <th>Omschrijving</th>
          <th class="num">Saldo</th>
        </tr>
      </thead>
      <tbody>
        {loan_rows}
        <tr style="font-weight:700; background:#eef2f8;">
          <td colspan="2">Totaal hypotheekschuld</td>
          <td class="num neg">{debt_total_fmt}</td>
        </tr>
        <tr style="background:#fff8f0;">
          <td colspan="2">Rentelasten Q1 (betaald)</td>
          <td class="num warn">{interest_fmt}</td>
        </tr>
      </tbody>
    </table>
  </div>

  <!-- Trial Balance summary -->
  <div class="section">
    <div class="section-title">Trial Balance – Selectie per rekening</div>
    <table>
      <thead>
        <tr>
          <th>Code</th>
          <th>Omschrijving</th>
          <th class="num">{tb_q1_label}</th>
          <th class="num">{tb_q2_label}</th>
          <th class="num">Mutatie</th>
        </tr>
      </thead>
      <tbody>
        {tb_rows}
      </tbody>
    </table>
  </div>

  <!-- BTW -->
  {btw_section}

</div>

<div class="footer">Rapport gegenereerd door Q2 Reporting Agent &nbsp;&bull;&nbsp; {generated}</div>
</body>
</html>
"""


def fmt_eur(value):
    if value is None:
        return "–"
    neg = value < 0
    v = abs(value)
    formatted = f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"({formatted})" if neg else formatted


def build_report(wb, output_path):
    print("Parsing sheets...")
    tb = parse_trial_balance(wb)
    banks = parse_banken(wb)
    ar = parse_ar(wb)
    ap = parse_ap(wb)
    rent = parse_periodieke_huur(wb)
    loans = parse_leningen(wb)
    deposits = parse_waarborgsommen(wb)
    tenancy = parse_tenancy_schedule(wb)
    btw = parse_btw(wb)

    company_name = "Newton Amsterdam B.V. (2200)"
    period = tb.get("q2_label") or "Q2 2026"
    generated = datetime.now().strftime("%d-%m-%Y %H:%M")

    # KPI values
    rent_q2_val = rent["q2_total"]
    bank_val = banks["total"]
    ar_val = ar["total_owed"]
    ap_val = ap["total"]
    debt_val = loans["total_debt"]
    deposit_val = deposits["total"]

    # Rent table
    months = rent["months"]
    rent_month_headers = "".join(f'<th class="num">{m}</th>' for m in months)

    month_totals = {m: 0.0 for m in months}
    rent_rows_html = []
    for t in sorted(rent["tenants"], key=lambda x: x["q2_total"], reverse=True):
        cells = ""
        for m in months:
            v = t["monthly"].get(m, 0.0)
            month_totals[m] = month_totals.get(m, 0.0) + v
            cells += f'<td class="num">{fmt_eur(v)}</td>'
        cls = "pos" if t["q2_total"] > 0 else "neg"
        rent_rows_html.append(
            f'<tr><td>{t["unit"]}</td><td>{t["tenant"]}</td><td style="font-size:11px;color:#8a9bb0">{t["description"]}</td>'
            f'{cells}<td class="num {cls}">{fmt_eur(t["q2_total"])}</td></tr>'
        )

    rent_total_months_html = "".join(
        f'<td class="num">{fmt_eur(month_totals[m])}</td>' for m in months
    )

    # AR rows
    ar_rows_html = []
    for r in sorted(ar["rows"], key=lambda x: abs(x["total_owed"]), reverse=True)[:20]:
        cls = "neg" if r["total_owed"] > 0 else "pos"
        ar_rows_html.append(
            f'<tr><td>{r["tenant"]}</td>'
            f'<td class="num">{fmt_eur(r["current_owed"])}</td>'
            f'<td class="num {cls}">{fmt_eur(r["total_owed"])}</td></tr>'
        )
    if not ar_rows_html:
        ar_rows_html = ['<tr><td colspan="3" style="text-align:center;color:#aaa">Geen openstaande debiteuren</td></tr>']

    # AP rows
    ap_rows_html = []
    for r in ap["rows"]:
        ap_rows_html.append(
            f'<tr><td>{r["name"]}</td>'
            f'<td class="num">{fmt_eur(r["current"])}</td>'
            f'<td class="num neg">{fmt_eur(r["total"])}</td></tr>'
        )
    if not ap_rows_html:
        ap_rows_html = ['<tr><td colspan="3" style="text-align:center;color:#aaa">Geen openstaande crediteuren</td></tr>']

    # Bank rows
    bank_rows_html = []
    for b in banks["banks"]:
        bank_rows_html.append(
            f'<tr><td>{b["code"]}</td><td>{b["name"]}</td>'
            f'<td class="num pos">{fmt_eur(b["balance"])}</td></tr>'
        )

    # Loan rows
    loan_rows_html = []
    for l in loans["loans"]:
        loan_rows_html.append(
            f'<tr><td>{l["code"]}</td><td>{l["name"]}</td>'
            f'<td class="num neg">{fmt_eur(l["balance"])}</td></tr>'
        )

    # Trial Balance – show accounts with significant movement
    tb_rows_html = []
    for acc in tb["accounts"]:
        mutatie = acc["q2_ending"] - acc["q1_ending"]
        if abs(acc["q2_ending"]) < 100 and abs(mutatie) < 100:
            continue
        cls = "pos" if mutatie > 0 else ("neg" if mutatie < 0 else "")
        tb_rows_html.append(
            f'<tr><td style="font-size:12px;color:#8a9bb0">{acc["code"]}</td>'
            f'<td>{acc["name"]}</td>'
            f'<td class="num">{fmt_eur(acc["q1_ending"])}</td>'
            f'<td class="num">{fmt_eur(acc["q2_ending"])}</td>'
            f'<td class="num {cls}">{fmt_eur(mutatie)}</td></tr>'
        )

    # BTW section
    if btw["q2"]:
        btw_rows = "".join(
            f'<tr><td>{k}</td><td class="num">{fmt_eur(v)}</td></tr>'
            for k, v in btw["q2"].items()
        )
        btw_section_html = f"""
  <div class="section">
    <div class="section-title">BTW Specificatie Q2</div>
    <table>
      <thead><tr><th>Omschrijving</th><th class="num">Bedrag</th></tr></thead>
      <tbody>{btw_rows}</tbody>
    </table>
  </div>"""
    else:
        btw_section_html = ""

    html = REPORT_TEMPLATE.format(
        company=company_name,
        period=period,
        generated=generated,
        rent_q2=fmt_eur(rent_q2_val),
        bank_total=fmt_eur(bank_val),
        ar_total=fmt_eur(ar_val),
        ap_total=fmt_eur(ap_val),
        debt_total=fmt_eur(debt_val),
        deposit_total=fmt_eur(deposit_val),
        rent_month_headers=rent_month_headers,
        rent_rows="\n".join(rent_rows_html),
        rent_total_months=rent_total_months_html,
        rent_q2_fmt=fmt_eur(rent_q2_val),
        ar_rows="\n".join(ar_rows_html),
        ap_rows="\n".join(ap_rows_html),
        bank_rows="\n".join(bank_rows_html),
        bank_total_fmt=fmt_eur(bank_val),
        loan_rows="\n".join(loan_rows_html),
        debt_total_fmt=fmt_eur(debt_val),
        interest_fmt=fmt_eur(loans["interest_total"]),
        tb_q1_label=tb.get("q1_label") or "Q1",
        tb_q2_label=tb.get("q2_label") or "Q2",
        tb_rows="\n".join(tb_rows_html),
        btw_section=btw_section_html,
    )

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Report written to: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Q2 Reporting Agent")
    parser.add_argument("excel_file", help="Path to the quarterly Excel file")
    parser.add_argument("--output", default="output", help="Output directory (default: output/)")
    args = parser.parse_args()

    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        print(f"Error: file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = excel_path.stem.replace(" ", "_")
    output_path = output_dir / f"{stem}_Q2_rapport.html"

    print(f"Loading: {excel_path.name}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    build_report(wb, output_path)
    print("Done.")
    return str(output_path)


if __name__ == "__main__":
    main()
